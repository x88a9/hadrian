"""Parser for the Hadrian_Engine result exports (Phase 5, T3/D2).

Pure and DB-free: turns the ``results/`` directory of Hadrian_Engine (one
sub-folder per system) into a list of
:class:`~app.importers.programmatic_types.ParsedProgrammaticSystem`. Never
raises for a per-system problem — a broken workbook / missing backtest file
becomes a ``skipped``/``incomplete`` system carrying a diagnostic ``message``.

Per-system layout (verified against the real data set)::

    results/<system>/
        results.xlsx        one row per run (Run#, System, Config Label, TF,
                            Symbol, Mode, n, WR%, EV(R), Sharpe, Calmar,
                            MaxDD(R), ProfitFactor, R/yr, [Trades/yr, Notes])
        backtest_<Run#>.xlsx  sheet 'Trades': rows 1-3 = Entry/Stop/TP rule
                            key-value, row 5 = header, rows 6.. = trades

Best-config selection (see docs/DECISIONS.md, "Best-config selection for engine
imports"): among rows deduplicated
on (Run#, Mode) with ``n >= 20`` and a numeric EV(R), restricted to backtest
modes (``full`` | ``IS``), pick the row with the **maximum EV(R)**, breaking ties
by preferring ``full`` and then the lowest Run#. Preferring ``full`` outright would select a
negative-EV auxiliary profiling run over the genuinely best configuration on at
least one real system; EV-maximisation with ``full`` only as a tiebreak
reproduces the expected best config everywhere while still preferring ``full``
runs where they are genuinely best.

Trades come from ``backtest_<Run#>.xlsx``; if that file is missing the system is
kept as reported-only (``incomplete``).
"""

from __future__ import annotations

import os
from datetime import datetime
from typing import Optional

import openpyxl

from app.importers.assets import derive_asset_from_symbol
from app.importers.programmatic_types import (
    ParsedProgrammaticSystem,
    ParsedTrade,
)
from app.services.metrics import derive_win_loss

# Result directories that are not systems: engine test scaffolds, and duplicate
# backups (any name ending in ``_backup``). Extend for a specific engine layout
# with HADRIAN_ENGINE_EXCLUDE, a comma-separated list of directory names.
_EXCLUDE = {"engine_test", "minimal_test"} | {
    name.strip()
    for name in os.environ.get("HADRIAN_ENGINE_EXCLUDE", "").split(",")
    if name.strip()
}
_EXCLUDE_SUFFIX = "_backup"

_MIN_N = 20
_BACKTEST_MODES = ("full", "IS")

_NAN = {"", "nan", "none", "null", "#n/a", "#div/0!", "#ref!", "#value!", "—"}

_DATETIME_FORMATS = (
    "%Y-%m-%d %H:%M:%S",
    "%Y-%m-%d %H:%M",
    "%Y-%m-%dT%H:%M:%S",
    "%Y-%m-%d",
)


# --------------------------------------------------------------------------- #
# Cell helpers
# --------------------------------------------------------------------------- #
def _clean_str(value) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if s == "" or s.lower() in _NAN:
        return None
    return s


def _to_float(value) -> Optional[float]:
    if value is None or isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        f = float(value)
    else:
        s = str(value).strip()
        if s.lower() in _NAN:
            return None
        try:
            f = float(s)
        except (ValueError, TypeError):
            return None
    if f != f or f in (float("inf"), float("-inf")):
        return None
    return f


def _to_int(value) -> Optional[int]:
    f = _to_float(value)
    return int(round(f)) if f is not None else None


def _to_datetime(date_val, time_val) -> Optional[datetime]:
    """Combine the ``Date`` and ``Entry Time`` cells into one datetime.

    ``Entry Time`` is usually already a full timestamp; fall back to the date
    cell (at 00:00) when only a date is available.
    """
    for candidate in (time_val, date_val):
        if isinstance(candidate, datetime):
            return candidate
    for candidate in (time_val, date_val):
        s = _clean_str(candidate)
        if s is None:
            continue
        try:
            return datetime.fromisoformat(s)
        except ValueError:
            pass
        for fmt in _DATETIME_FORMATS:
            try:
                return datetime.strptime(s, fmt)
            except ValueError:
                continue
    return None


def _norm_direction(value) -> Optional[str]:
    s = _clean_str(value)
    if s is None:
        return None
    s = s.lower()
    if s in ("long", "buy"):
        return "long"
    if s in ("short", "sell"):
        return "short"
    return None


def _norm_win_loss(value, r: Optional[float]) -> Optional[str]:
    s = _clean_str(value)
    if s is not None:
        s = s.lower()
        if s in ("w", "win"):
            return "win"
        if s in ("l", "loss"):
            return "loss"
    return derive_win_loss(r)


# --------------------------------------------------------------------------- #
# results.xlsx -> run rows
# --------------------------------------------------------------------------- #
def _read_runs(path: str) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return []

    header = [(_clean_str(c) or "") for c in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    def cell(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    runs: list[dict] = []
    for row in rows[1:]:
        if not row or row[idx.get("Run#", 0)] is None:
            continue
        runs.append(
            {
                "run": _to_int(cell(row, "Run#")),
                "system": _clean_str(cell(row, "System")),
                "config_label": _clean_str(cell(row, "Config Label")),
                "tf": _clean_str(cell(row, "TF")),
                "symbol": _clean_str(cell(row, "Symbol")),
                "mode": _clean_str(cell(row, "Mode")),
                "n": _to_int(cell(row, "n")),
                "wr": _to_float(cell(row, "WR%")),
                "ev": _to_float(cell(row, "EV(R)")),
                "sharpe": _to_float(cell(row, "Sharpe")),
                "calmar": _to_float(cell(row, "Calmar")),
                "max_dd": _to_float(cell(row, "MaxDD(R)")),
                "profit_factor": _to_float(cell(row, "ProfitFactor")),
                "r_per_year": _to_float(cell(row, "R/yr")),
                "trades_per_year": _to_float(cell(row, "Trades/yr")),
            }
        )
    return runs


def _dedup_runs(runs: list[dict]) -> list[dict]:
    seen: dict[tuple, dict] = {}
    for r in runs:
        key = (r["run"], r["mode"])
        seen.setdefault(key, r)
    return list(seen.values())


def _select_best(runs: list[dict]) -> Optional[dict]:
    eligible = [
        r
        for r in runs
        if r["run"] is not None
        and r["n"] is not None
        and r["n"] >= _MIN_N
        and r["ev"] is not None
        and r["mode"] in _BACKTEST_MODES
    ]
    if not eligible:
        return None
    # Max EV(R); tiebreak prefer 'full', then lowest Run#.
    return max(eligible, key=lambda r: (r["ev"], r["mode"] == "full", -r["run"]))


def _companion_ev(runs: list[dict], best: dict, mode: str) -> Optional[float]:
    """EV of the IS/OOS run sharing config label (+TF+Symbol) with ``best``.

    IS and OOS live in separate rows with different Run#, so they are matched by
    config identity rather than Run#.
    """
    for r in runs:
        if (
            r["mode"] == mode
            and r["config_label"] == best["config_label"]
            and r["tf"] == best["tf"]
            and r["symbol"] == best["symbol"]
        ):
            return r["ev"]
    return None


def _reported_metrics(runs: list[dict], best: dict) -> dict:
    return {
        "total_trades": best["n"],
        "win_rate": (best["wr"] / 100.0) if best["wr"] is not None else None,
        "ev": best["ev"],
        "sharpe": best["sharpe"],
        "calmar": best["calmar"],
        "max_drawdown_r": best["max_dd"],
        "profit_factor": best["profit_factor"],
        "r_per_year": best["r_per_year"],
        "trades_per_year": best["trades_per_year"],
        "is_ev": _companion_ev(runs, best, "IS"),
        "oos_ev": _companion_ev(runs, best, "OOS"),
        "config_label": best["config_label"],
        "run_number": best["run"],
    }


# --------------------------------------------------------------------------- #
# backtest_<Run#>.xlsx -> rules + trades
# --------------------------------------------------------------------------- #
def _read_backtest(path: str, timeframe: Optional[str]):
    """Return (entry_rule, sl_rule, tp_rule, trades) from a backtest workbook."""
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb["Trades"] if "Trades" in wb.sheetnames else wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()

    def rule(i: int) -> Optional[str]:
        return _clean_str(rows[i][1]) if len(rows) > i and len(rows[i]) > 1 else None

    entry_rule = rule(0)
    sl_rule = rule(1)
    tp_rule = rule(2)

    # Find header row (the one whose first cell is '#').
    header_row = None
    for i, row in enumerate(rows):
        if row and _clean_str(row[0]) == "#":
            header_row = i
            break
    if header_row is None:
        return entry_rule, sl_rule, tp_rule, []

    header = [(_clean_str(c) or "") for c in rows[header_row]]
    idx = {name: i for i, name in enumerate(header)}

    def cell(row, name):
        i = idx.get(name)
        return row[i] if i is not None and i < len(row) else None

    trades: list[ParsedTrade] = []
    for row in rows[header_row + 1 :]:
        if not row or row[0] is None:
            continue
        r_value = _to_float(cell(row, "R_net"))
        trades.append(
            ParsedTrade(
                trade_datetime=_to_datetime(cell(row, "Date"), cell(row, "Entry Time")),
                timeframe=timeframe,
                entry=_to_float(cell(row, "Entry$")),
                sl=_to_float(cell(row, "SL$")),
                exit=_to_float(cell(row, "Exit$")),
                direction=_norm_direction(cell(row, "Direction")),
                r_value=r_value,
                win_loss=_norm_win_loss(cell(row, "W/L"), r_value),
            )
        )
    return entry_rule, sl_rule, tp_rule, trades


# --------------------------------------------------------------------------- #
# Public API
# --------------------------------------------------------------------------- #
def _parse_system(system_dir: str, name: str) -> ParsedProgrammaticSystem:
    results_path = os.path.join(system_dir, "results.xlsx")
    system = ParsedProgrammaticSystem(name=name, source_engine="hadrian_engine")

    if not os.path.isfile(results_path):
        system.parse_status = "skipped"
        system.message = "no results.xlsx"
        return system

    try:
        runs = _dedup_runs(_read_runs(results_path))
    except Exception as exc:  # noqa: BLE001 - never crash the whole run
        system.parse_status = "skipped"
        system.message = f"results.xlsx parse error: {exc!r}"
        return system

    best = _select_best(runs)
    if best is None:
        system.parse_status = "incomplete"
        system.message = "no eligible best-config run (n>=20)"
        return system

    system.timeframe = best["tf"]
    # Asset: the symbol of EXACTLY the selected best-config row, the same row
    # the metrics and trades come from. No separate selection rule.
    system.asset = derive_asset_from_symbol(best["symbol"], context=f"engine:{name}")
    system.reported_metrics = _reported_metrics(runs, best)

    backtest_path = os.path.join(system_dir, f"backtest_{best['run']}.xlsx")
    if not os.path.isfile(backtest_path):
        system.parse_status = "incomplete"
        system.message = f"backtest_{best['run']}.xlsx missing (reported-only)"
        return system

    try:
        entry_rule, sl_rule, tp_rule, trades = _read_backtest(
            backtest_path, best["tf"]
        )
    except Exception as exc:  # noqa: BLE001
        system.parse_status = "incomplete"
        system.message = f"backtest parse error: {exc!r}"
        return system

    system.entry_rule = entry_rule
    system.sl_rule = sl_rule
    system.tp_rule = tp_rule
    system.trades = trades
    has_r = any(t.r_value is not None for t in trades)
    system.parse_status = "complete" if has_r else "incomplete"
    return system


def parse_hadrian_engine(results_dir: str) -> list[ParsedProgrammaticSystem]:
    """Parse a Hadrian_Engine ``results`` directory into programmatic systems.

    Missing directory -> ``[]``. Excluded folders (test scaffolds / backups) are
    skipped. Never raises for a per-system problem.
    """
    if not results_dir or not os.path.isdir(results_dir):
        return []

    systems: list[ParsedProgrammaticSystem] = []
    for name in sorted(os.listdir(results_dir)):
        system_dir = os.path.join(results_dir, name)
        if (
            not os.path.isdir(system_dir)
            or name in _EXCLUDE
            or name.endswith(_EXCLUDE_SUFFIX)
        ):
            continue
        try:
            systems.append(_parse_system(system_dir, name))
        except Exception as exc:  # noqa: BLE001 - never crash the whole run
            systems.append(
                ParsedProgrammaticSystem(
                    name=name,
                    source_engine="hadrian_engine",
                    parse_status="skipped",
                    message=f"unexpected error: {exc!r}",
                )
            )
    return systems
