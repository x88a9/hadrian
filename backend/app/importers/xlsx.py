"""Header-driven parser for the canonical ``Backtesting Repository.xlsx``.

Pure, DB-free. Turns a workbook into a :class:`ParseResult` of
:class:`ParsedTab` objects that the import service (app/services/import_service)
persists. Robustness is a hard requirement: the workbook mixes two tab layouts
and contains unfinished backtests full of Excel error strings
(``#DIV/0!``/``#N/A``/…). The parser must never crash the whole run — any
unexpected per-tab error is caught and turned into a ``skipped`` tab result.

Layout variants (verified, see docs/DECISIONS.md, "Two tab layout variants"):

- Variant A (majority): system name in ``B2``; metric header row 7, values
  row 9; trade header ``# | Day | Date & Time | Zone | Timeframe | Entry ($) |
  Stop Loss ($) | Exit ($) | (blank) | Direction | R | W/L``.
- Variant B (older tabs): system name in ``B3``; metric values row 8; trade
  header uses ``Duration | Date Start | Date End`` and Direction/R/W-L shifted.

Parsing is therefore driven by header *text*, not fixed coordinates.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from datetime import datetime
from typing import Optional

import openpyxl

from app.importers.assets import derive_asset_from_timeframe_cells

# Excel error sentinels that mean "not computable" -> None.
_ERROR_STRINGS = {
    "#DIV/0!",
    "#N/A",
    "#REF!",
    "#VALUE!",
    "#NAME?",
    "#NULL!",
    "#NUM!",
}

# Metric header label (normalized) -> reported_metrics key.
_METRIC_HEADERS = {
    "composite grade": "composite_grade",
    "ev grade": "ev_grade",
    "ece grade": "ece_grade",
    "evol grade": "evol_grade",
    "win rate": "win_rate",
    "ev": "ev",
    "total r": "total_r",
    "avg win r": "avg_win_r",
    "avg loss r": "avg_loss_r",
    "total trades": "total_trades",
    "wins": "wins",
    "losses": "losses",
    "ece": "ece",
    "evol": "evol",
}

_NUMERIC_METRIC_KEYS = {
    "win_rate",
    "ev",
    "total_r",
    "avg_win_r",
    "avg_loss_r",
    "total_trades",
    "wins",
    "losses",
    "ece",
    "evol",
}

_GRADE_METRIC_KEYS = {
    "composite_grade",
    "ev_grade",
    "ece_grade",
    "evol_grade",
}

# Rows scanned for the name / rule / metric-header blocks.
_HEADER_SCAN_ROWS = 15
_RULE_SCAN_ROWS = 10


@dataclass
class ParsedTrade:
    number: Optional[float] = None
    day: Optional[str] = None
    trade_datetime: Optional[datetime] = None
    zone: Optional[str] = None
    timeframe: Optional[str] = None
    entry: Optional[float] = None
    sl: Optional[float] = None
    exit: Optional[float] = None
    direction: Optional[str] = None  # "long" | "short" | None
    r_value: Optional[float] = None
    win_loss: Optional[str] = None  # "win" | "loss" | "draw" | None


@dataclass
class ParsedTab:
    tab_name: str
    system_name: str
    # Traded asset, set only where the tab evidences one. None means no
    # evidence; the import then applies the default (app/importers/assets.py).
    asset: Optional[str] = None
    entry_rule: Optional[str] = None
    sl_rule: Optional[str] = None
    tp_rule: Optional[str] = None
    reported_metrics: Optional[dict] = None
    trades: list[ParsedTrade] = field(default_factory=list)
    parse_status: str = "incomplete"  # "complete" | "incomplete" | "skipped"
    message: Optional[str] = None


@dataclass
class ParseResult:
    tabs: list[ParsedTab] = field(default_factory=list)


# --------------------------------------------------------------------------- #
# Cell-value helpers
# --------------------------------------------------------------------------- #
def _clean(value):
    """Normalize a raw cell value: Excel errors / blanks / "None" -> None."""
    if value is None:
        return None
    if isinstance(value, str):
        s = value.strip()
        if s == "" or s in _ERROR_STRINGS or s.lower() == "none":
            return None
        return s
    return value


def _as_str(value) -> Optional[str]:
    v = _clean(value)
    if v is None:
        return None
    return str(v).strip() or None


def _as_number(value) -> Optional[float]:
    v = _clean(value)
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return v
    return None


def _as_datetime(value) -> Optional[datetime]:
    # Only genuine datetime instances count; date strings are ignored.
    if isinstance(value, datetime):
        return value
    return None


def _fallback_win_loss(r: Optional[float]) -> Optional[str]:
    """xlsx W/L fallback from R (kept local to keep the parser dependency-free).

    R > 0 -> win, R < -0.1 -> loss, R == 0 -> draw, else None.
    """
    if r is None:
        return None
    if r > 0:
        return "win"
    if r < -0.1:
        return "loss"
    if r == 0:
        return "draw"
    return None


def _normalize_direction(value) -> Optional[str]:
    v = _as_str(value)
    if v is None:
        return None
    v = v.lower()
    return v if v in ("long", "short") else None


def _normalize_win_loss(value, r: Optional[float]) -> Optional[str]:
    v = _as_str(value)
    if v is not None:
        v = v.lower()
        if v in ("win", "loss", "draw"):
            return v
    return _fallback_win_loss(r)


# --------------------------------------------------------------------------- #
# Grid access
# --------------------------------------------------------------------------- #
def _cell(grid: list[tuple], row: int, col: int):
    """1-based (row, col) access on a materialized value grid; None if OOB."""
    if 1 <= row <= len(grid):
        r = grid[row - 1]
        if 1 <= col <= len(r):
            return r[col - 1]
    return None


# --------------------------------------------------------------------------- #
# Block parsers
# --------------------------------------------------------------------------- #
def _resolve_name(grid: list[tuple], tab_name: str) -> str:
    return _as_str(_cell(grid, 2, 2)) or _as_str(_cell(grid, 3, 2)) or tab_name.strip()


def _parse_rules(grid: list[tuple]) -> tuple[Optional[str], Optional[str], Optional[str]]:
    labels = {"entry": None, "stop loss": None, "take profit": None}
    max_row = min(len(grid), _RULE_SCAN_ROWS)
    for row in range(1, max_row + 1):
        r = grid[row - 1]
        for col in range(1, len(r) + 1):
            label = _as_str(r[col - 1])
            if label is None:
                continue
            key = label.lower()
            if key in labels and labels[key] is None:
                labels[key] = _as_str(_cell(grid, row, col + 1))
    return labels["entry"], labels["stop loss"], labels["take profit"]


def _parse_reported_metrics(grid: list[tuple]) -> Optional[dict]:
    # Find the metric header row (the one carrying "Composite Grade").
    header_row = None
    composite_col = None
    max_row = min(len(grid), _HEADER_SCAN_ROWS)
    for row in range(1, max_row + 1):
        r = grid[row - 1]
        for col in range(1, len(r) + 1):
            label = _as_str(r[col - 1])
            if label is not None and label.lower() == "composite grade":
                header_row = row
                composite_col = col
                break
        if header_row is not None:
            break

    if header_row is None:
        return None

    # Map every recognized metric header to its column.
    header_cols: dict[str, int] = {}
    hr = grid[header_row - 1]
    for col in range(1, len(hr) + 1):
        label = _as_str(hr[col - 1])
        if label is None:
            continue
        key = _METRIC_HEADERS.get(label.lower())
        if key is not None and key not in header_cols:
            header_cols[key] = col

    # Value row = first row below the header whose Composite-Grade column is
    # non-empty (variant A -> row 9, variant B -> row 8). Fallback: first row
    # where any mapped metric column is non-empty.
    value_row = None
    for row in range(header_row + 1, header_row + 6):
        if _clean(_cell(grid, row, composite_col)) is not None:
            value_row = row
            break
    if value_row is None:
        for row in range(header_row + 1, header_row + 6):
            if any(_clean(_cell(grid, row, c)) is not None for c in header_cols.values()):
                value_row = row
                break
    if value_row is None:
        value_row = header_row + 1

    metrics: dict = {}
    for key in _METRIC_HEADERS.values():
        col = header_cols.get(key)
        raw = _cell(grid, value_row, col) if col is not None else None
        if key in _NUMERIC_METRIC_KEYS:
            metrics[key] = _as_number(raw)
        else:  # grade key
            metrics[key] = _as_str(raw)
    return metrics


def _find_trade_header_row(grid: list[tuple]) -> Optional[int]:
    for row in range(1, len(grid) + 1):
        if _as_str(_cell(grid, row, 1)) == "#":
            return row
    return None


def _map_trade_columns(grid: list[tuple], header_row: int) -> dict[int, str]:
    """header column (1-based) -> ParsedTrade field name."""
    mapping: dict[int, str] = {}
    hr = grid[header_row - 1]
    for col in range(1, len(hr) + 1):
        label = _as_str(hr[col - 1])
        if label is None:
            continue
        low = label.lower()
        if low == "#":
            field_name = "number"
        elif low == "day":
            field_name = "day"
        elif low == "date end" or low == "duration":
            continue  # explicitly ignored
        elif low.startswith("date"):  # "date & time" / "date start" / "date"
            field_name = "trade_datetime"
        elif low == "zone":
            field_name = "zone"
        elif low == "timeframe":
            field_name = "timeframe"
        elif low.startswith("entry"):
            field_name = "entry"
        elif low.startswith("stop loss"):
            field_name = "sl"
        elif low.startswith("exit"):
            field_name = "exit"
        elif low == "direction":
            field_name = "direction"
        elif low == "r":
            field_name = "r_value"
        elif low == "w/l":
            field_name = "win_loss"
        else:
            continue
        mapping.setdefault(col, field_name)
    return mapping


def _parse_trades(grid: list[tuple], header_row: int, colmap: dict[int, str]) -> list[ParsedTrade]:
    # invert: field -> column
    field_col = {v: k for k, v in colmap.items()}
    entry_col = field_col.get("entry")
    trades: list[ParsedTrade] = []
    for row in range(header_row + 1, len(grid) + 1):
        # A row is a trade iff the entry price is non-empty (xlsx convention).
        entry = _as_number(_cell(grid, row, entry_col)) if entry_col is not None else None
        if entry is None:
            continue

        r_value = _as_number(_cell(grid, row, field_col["r_value"])) if "r_value" in field_col else None

        trade = ParsedTrade(
            number=_as_number(_cell(grid, row, field_col["number"])) if "number" in field_col else None,
            day=_as_str(_cell(grid, row, field_col["day"])) if "day" in field_col else None,
            trade_datetime=_as_datetime(_cell(grid, row, field_col["trade_datetime"]))
            if "trade_datetime" in field_col
            else None,
            zone=_as_str(_cell(grid, row, field_col["zone"])) if "zone" in field_col else None,
            timeframe=_as_str(_cell(grid, row, field_col["timeframe"])) if "timeframe" in field_col else None,
            entry=entry,
            sl=_as_number(_cell(grid, row, field_col["sl"])) if "sl" in field_col else None,
            exit=_as_number(_cell(grid, row, field_col["exit"])) if "exit" in field_col else None,
            direction=_normalize_direction(_cell(grid, row, field_col["direction"]))
            if "direction" in field_col
            else None,
            r_value=r_value,
            win_loss=_normalize_win_loss(
                _cell(grid, row, field_col["win_loss"]) if "win_loss" in field_col else None,
                r_value,
            ),
        )
        trades.append(trade)
    return trades


def _parse_tab(grid: list[tuple], tab_name: str) -> ParsedTab:
    system_name = _resolve_name(grid, tab_name)
    tab = ParsedTab(tab_name=tab_name, system_name=system_name)

    tab.entry_rule, tab.sl_rule, tab.tp_rule = _parse_rules(grid)
    tab.reported_metrics = _parse_reported_metrics(grid)

    header_row = _find_trade_header_row(grid)
    if header_row is None:
        tab.parse_status = "skipped"
        tab.message = "no trade header row (no '#' cell in column A)"
        return tab

    colmap = _map_trade_columns(grid, header_row)
    if "entry" not in colmap.values():
        tab.parse_status = "skipped"
        tab.message = "trade header row has no 'Entry ($)' column"
        return tab

    tab.trades = _parse_trades(grid, header_row, colmap)
    # Asset: some tabs carry the ticker in the "Timeframe" column instead of a
    # timeframe. Only evidenced matches are set.
    tab.asset = derive_asset_from_timeframe_cells(
        (t.timeframe for t in tab.trades), context=f"xlsx:{tab_name}"
    )
    has_numeric_r = any(t.r_value is not None for t in tab.trades)
    tab.parse_status = "complete" if has_numeric_r else "incomplete"
    return tab


def parse_workbook(path: str) -> ParseResult:
    """Parse every tab of the workbook into a :class:`ParseResult`.

    Never raises for per-tab problems: a broken tab becomes a ``skipped``
    :class:`ParsedTab` carrying a diagnostic ``message``.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    result = ParseResult()
    try:
        for tab_name in wb.sheetnames:
            try:
                ws = wb[tab_name]
                grid = list(ws.iter_rows(values_only=True))
                result.tabs.append(_parse_tab(grid, tab_name))
            except Exception as exc:  # noqa: BLE001 - never crash the whole run
                result.tabs.append(
                    ParsedTab(
                        tab_name=tab_name,
                        system_name=tab_name.strip(),
                        parse_status="skipped",
                        message=f"unexpected parse error: {exc!r}",
                    )
                )
    finally:
        wb.close()
    return result
