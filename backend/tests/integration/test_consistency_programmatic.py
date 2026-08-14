"""Phase-5 consistency gate (T7 / D7).

Runs the *real* programmatic import (Hadrian² + Hadrian_Engine) into the
``hadrian3_test`` database and reconciles the persisted systems / trades /
sweeps against the original Hadrian² source files. Auto-skipped when either
source directory is missing or the Postgres test server is unreachable
(pattern mirrored from ``test_reconciliation.py`` / ``test_import_programmatic``).

The five checks and their tolerances follow docs/DECISIONS.md,
"Reconciliation tolerances", with two binding corrections:

1. ``audit_master.csv`` metrics for the three audited systems MR-M15-101,
   MR-M15-101.02, MR-H1-101 — computed from the persisted trades via the real
   metric service, matched to the source row by
   ``reported_metrics['source_variant_id'] == variant_id``. Absolute tol 1.5e-4
   (the CSV stores four decimals); IS/OOS cutoff is identical (2024-01-01), so
   the split is exactly comparable. ``total_trades`` is exact.
2. Sweep exactness — every stored ``points['value']`` equals the source
   ``oos_net_ev`` bit-for-bit (no rounding), verified as a full per-carrier
   multiset plus three explicitly named sample cells (B-H1-101, MR-H1-102,
   MR-H4-101). Global counts: 32 programmatic systems, 124 sweeps, 2284 points.
3. Monte-Carlo tolerance band on the **OOS subset** (corrected): the source
   ``mc_*`` columns were computed over the OOS trades only (``mc_n_oos``), so we
   bootstrap ``quant.monte_carlo`` over the trades with
   ``trade_datetime >= 2024-01-01``. ``ev_p50`` within ±0.08 of ``mc_ev_p50``,
   ``p_ev_positive`` within ±0.10 of ``mc_pct_pos/100``; ``len(oos_r) ==
   mc_n_oos`` exactly. Statistical band (different RNG), not bit-equality.
4. Walk-forward plausibility — ``quant.walk_forward`` (defaults 6/3) over the
   dated trades: ``pct_positive*100`` within ±20 pp of ``wf_pct_positive`` and
   ``n_windows >= 10``. Plausibility, not equality (the source window cut is not
   identically reconstructible).
5. Engine reconciliation: for the system named by
   ``HADRIAN3_ENGINE_RECON`` (``<system>:<config>:<n>``), the persisted
   ``metrics.all.total_trades`` matches ``<n>`` and ``reported_metrics['ev']``
   equals that config's EV(R) read live from its ``results.xlsx``.
"""

from __future__ import annotations

import csv
import os
from collections import Counter
from datetime import datetime
from pathlib import Path

import openpyxl
import pytest
from sqlalchemy import select
from sqlalchemy.orm import sessionmaker

from app.core.config import settings
from app.models import ParameterSweep, System, Trade
from app.services.import_service import run_programmatic_import
from app.services.metrics import compute_all
from app.services import quant

pytestmark = pytest.mark.integration

from tests.paths import ENGINE_DIR, HADRIAN2_DIR, has_engine_sources

AUDIT_MASTER = Path(HADRIAN2_DIR) / "audit" / "audit_master.csv"

skip_no_sources = pytest.mark.skipif(
    not has_engine_sources(),
    reason="programmatic source dirs unset (HADRIAN2_RESULTS_DIR / "
    "HADRIAN_ENGINE_RESULTS_DIR)",
)

# Absolute tolerances (D7).
METRIC_ABS_TOL = 1.5e-4
MC_P50_TOL = 0.08
MC_PPOS_TOL = 0.10
WF_PCT_TOL = 20.0  # percentage points

AUDITED = ("MR-M15-101", "MR-M15-101.02", "MR-H1-101")
CARRIERS = {
    "B-H1-101": ("system1_results.csv", None),
    "MR-H1-102": ("system2_results.csv", None),
    "MR-H4-101": ("system3_results.csv", "4h"),
}

_NAN = {"", "nan", "none", "null", "#n/a", "#div/0!", "#ref!", "#value!"}


def _to_float(value):
    """Same lenient float parse the parser uses (errors/blank -> None)."""
    if value is None:
        return None
    s = str(value).strip()
    if s.lower() in _NAN:
        return None
    try:
        f = float(s)
    except (ValueError, TypeError):
        return None
    if f != f or f in (float("inf"), float("-inf")):
        return None
    return f


# --------------------------------------------------------------------------- #
# Module-scoped real import (heavy — run once, shared by all checks).
# --------------------------------------------------------------------------- #
@pytest.fixture(scope="module")
def imported(db_engine, session_factory):
    """Truncate the test DB once and run the real programmatic import."""
    from sqlalchemy import text

    with db_engine.begin() as conn:
        tables = conn.execute(
            text(
                "SELECT tablename FROM pg_tables "
                "WHERE schemaname = 'public' AND tablename <> 'alembic_version'"
            )
        ).scalars().all()
        if tables:
            joined = ", ".join(f'"{t}"' for t in tables)
            conn.execute(text(f"TRUNCATE {joined} RESTART IDENTITY CASCADE"))

    Session = sessionmaker(bind=db_engine, autoflush=False, future=True)
    session = Session()
    try:
        run_programmatic_import(session, HADRIAN2_DIR, ENGINE_DIR)
        session.commit()
        yield session
    finally:
        session.close()


def _system(session, name: str) -> System:
    return session.execute(
        select(System).where(System.name == name)
    ).scalar_one()


def _audit_by_variant() -> dict[str, dict]:
    with open(AUDIT_MASTER, encoding="utf-8-sig", newline="") as fh:
        return {row["variant_id"]: row for row in csv.DictReader(fh)}


def _csv_rows(filename: str, timeframe: str | None) -> list[dict]:
    with open(Path(HADRIAN2_DIR) / filename, encoding="utf-8-sig", newline="") as fh:
        rows = list(csv.DictReader(fh))
    if timeframe is not None:
        rows = [r for r in rows if (r.get("timeframe") or "").strip() == timeframe]
    return rows


# --------------------------------------------------------------------------- #
# Test 1 — audit_master metric reconciliation.
# --------------------------------------------------------------------------- #
@skip_no_sources
def test_audit_master_metrics(imported):
    audit = _audit_by_variant()
    split = settings.IS_OOS_SPLIT_DATE
    report: list[str] = []
    failures: list[str] = []

    for name in AUDITED:
        sys = _system(imported, name)
        variant_id = sys.reported_metrics["source_variant_id"]
        assert variant_id in audit, f"{name}: variant_id {variant_id!r} not in audit_master"
        src = audit[variant_id]
        m = compute_all(list(sys.trades), split)

        # total_trades exact.
        n_src = int(round(_to_float(src["n_trades"])))
        if m["all"]["total_trades"] != n_src:
            failures.append(
                f"{name}: total_trades computed={m['all']['total_trades']} vs src={n_src} (exact)"
            )

        checks = [
            ("ev_all", m["all"]["ev"], _to_float(src["net_ev"])),
            ("ev_is", m["is"]["ev"], _to_float(src["is_net_ev"])),
            ("ev_oos", m["oos"]["ev"], _to_float(src["oos_net_ev"])),
            ("win_rate", m["all"]["win_rate"], _to_float(src["win_rate"])),
        ]
        for label, computed, src_val in checks:
            diff = abs(computed - src_val)
            report.append(f"{name}.{label}: computed={computed:.6f} src={src_val:.4f} |Δ|={diff:.2e}")
            if diff > METRIC_ABS_TOL:
                failures.append(
                    f"{name}.{label}: computed={computed!r} vs src={src_val!r} "
                    f"|Δ|={diff:.3e} > tol {METRIC_ABS_TOL:.1e}"
                )

    print("\n[T7 audit_master]\n  " + "\n  ".join(report))
    assert not failures, "audit_master divergence:\n  " + "\n  ".join(failures)


# --------------------------------------------------------------------------- #
# Test 2 — sweep exactness and global counts.
# --------------------------------------------------------------------------- #
@skip_no_sources
def test_sweep_exactness_and_counts(imported):
    prog = imported.execute(
        select(System).where(System.provenance == "programmatic")
    ).scalars().all()
    assert len(prog) == 32, f"expected 32 programmatic systems, got {len(prog)}"

    sweeps = imported.execute(select(ParameterSweep)).scalars().all()
    assert len(sweeps) == 124, f"expected 124 sweeps, got {len(sweeps)}"
    total_points = sum(len(s.points or []) for s in sweeps)
    assert total_points == 2284, f"expected 2284 sweep points, got {total_points}"

    def _value_counter(values) -> Counter:
        return Counter(
            (round(v, 12) if v is not None else None) for v in values
        )

    samples: list[str] = []
    for carrier, (filename, timeframe) in CARRIERS.items():
        sys = _system(imported, carrier)
        stored_sweeps = imported.execute(
            select(ParameterSweep).where(ParameterSweep.system_id == sys.id)
        ).scalars().all()
        stored_vals = [p["value"] for sw in stored_sweeps for p in (sw.points or [])]

        rows = _csv_rows(filename, timeframe)
        src_vals = [_to_float(r["oos_net_ev"]) for r in rows]

        # Full per-carrier multiset: every stored cell exactly equals a source
        # oos_net_ev (no rounding). This subsumes the "≥3 sample cells" clause.
        assert _value_counter(stored_vals) == _value_counter(src_vals), (
            f"{carrier}: stored sweep values do not match {filename} oos_net_ev multiset "
            f"(stored n={len(stored_vals)}, src n={len(src_vals)})"
        )

        # Explicit named sample cell: first source row (4h for system3), matched
        # to its stored point by (net_ev, n_trades), value compared bit-exact.
        first = rows[0]
        want_val = _to_float(first["oos_net_ev"])
        want_net = _to_float(first["net_ev"])
        want_n = int(round(_to_float(first["n_trades"])))
        match = next(
            p
            for sw in stored_sweeps
            for p in (sw.points or [])
            if p["net_ev"] == want_net and p["n_trades"] == want_n
        )
        assert match["value"] == want_val, (
            f"{carrier} sample cell (variant {first['variant_id']}): "
            f"stored value={match['value']!r} vs src oos_net_ev={want_val!r}"
        )
        samples.append(
            f"{carrier}: sample variant {first['variant_id']} value={match['value']!r} "
            f"== src {want_val!r}"
        )

    print("\n[T7 sweeps] 32 systems / 124 sweeps / 2284 points; exact per-carrier multiset OK\n  "
          + "\n  ".join(samples))


# --------------------------------------------------------------------------- #
# Test 3 — Monte-Carlo tolerance band on the OOS subset (corrected).
# --------------------------------------------------------------------------- #
@skip_no_sources
def test_monte_carlo_oos_band(imported):
    audit = _audit_by_variant()
    split_dt = datetime(
        settings.IS_OOS_SPLIT_DATE.year,
        settings.IS_OOS_SPLIT_DATE.month,
        settings.IS_OOS_SPLIT_DATE.day,
    )
    report: list[str] = []
    failures: list[str] = []

    for name in AUDITED:
        sys = _system(imported, name)
        src = audit[sys.reported_metrics["source_variant_id"]]

        oos_r = [
            t.r_value
            for t in sys.trades
            if t.trade_datetime is not None
            and t.trade_datetime >= split_dt
            and t.r_value is not None
        ]
        mc_n_oos = int(round(_to_float(src["mc_n_oos"])))
        if len(oos_r) != mc_n_oos:
            failures.append(f"{name}: len(oos_r)={len(oos_r)} vs mc_n_oos={mc_n_oos} (exact)")

        mc = quant.monte_carlo(oos_r, n_iterations=1000, seed=42)
        src_p50 = _to_float(src["mc_ev_p50"])
        src_ppos = _to_float(src["mc_pct_pos"]) / 100.0
        d_p50 = abs(mc["ev_p50"] - src_p50)
        d_ppos = abs(mc["p_ev_positive"] - src_ppos)
        report.append(
            f"{name}: oos_n={len(oos_r)} p50 comp={mc['ev_p50']:.4f} src={src_p50:.4f} "
            f"|Δ|={d_p50:.3f}; ppos comp={mc['p_ev_positive']:.3f} src={src_ppos:.3f} |Δ|={d_ppos:.3f}"
        )
        if d_p50 > MC_P50_TOL:
            failures.append(f"{name}: mc ev_p50 |Δ|={d_p50:.3f} > {MC_P50_TOL}")
        if d_ppos > MC_PPOS_TOL:
            failures.append(f"{name}: mc p_ev_positive |Δ|={d_ppos:.3f} > {MC_PPOS_TOL}")

    print("\n[T7 monte-carlo OOS]\n  " + "\n  ".join(report))
    assert not failures, "monte-carlo band violation:\n  " + "\n  ".join(failures)


# --------------------------------------------------------------------------- #
# Test 4 — Walk-forward plausibility.
# --------------------------------------------------------------------------- #
@skip_no_sources
def test_walk_forward_plausibility(imported):
    audit = _audit_by_variant()
    report: list[str] = []
    failures: list[str] = []

    for name in AUDITED:
        sys = _system(imported, name)
        src = audit[sys.reported_metrics["source_variant_id"]]
        wf = quant.walk_forward(list(sys.trades))
        comp_pct = (wf["pct_positive"] or 0.0) * 100.0
        src_pct = _to_float(src["wf_pct_positive"])
        diff = abs(comp_pct - src_pct)
        report.append(
            f"{name}: pct comp={comp_pct:.1f} src={src_pct:.1f} |Δ|={diff:.1f}pp "
            f"n_windows={wf['n_windows']}"
        )
        if diff > WF_PCT_TOL:
            failures.append(f"{name}: wf pct |Δ|={diff:.1f}pp > {WF_PCT_TOL}pp")
        if wf["n_windows"] < 10:
            failures.append(f"{name}: wf n_windows={wf['n_windows']} < 10")

    print("\n[T7 walk-forward]\n  " + "\n  ".join(report))
    assert not failures, "walk-forward plausibility violation:\n  " + "\n  ".join(failures)


# --------------------------------------------------------------------------- #
# Test 5 — Engine reconciliation against a configured system.
# --------------------------------------------------------------------------- #
# "<system>:<config>:<n>", e.g. exported from a local engine run. Unset -> skip,
# so no system identifier from a private engine layout lives in this repository.
_RECON = os.environ.get("HADRIAN3_ENGINE_RECON", "")


@skip_no_sources
@pytest.mark.skipif(not _RECON, reason="HADRIAN3_ENGINE_RECON unset")
def test_engine_reconciliation(imported):
    name, config, n_str = _RECON.split(":")
    n_expected = int(n_str)

    sys = _system(imported, name)
    m = compute_all(list(sys.trades), settings.IS_OOS_SPLIT_DATE)
    assert m["all"]["total_trades"] == n_expected, (
        f"{name} total_trades computed={m['all']['total_trades']} "
        f"vs expected {n_expected} (best config {config})"
    )

    # Read that config's EV(R) live from results.xlsx.
    results_path = Path(ENGINE_DIR) / name / "results.xlsx"
    wb = openpyxl.load_workbook(results_path, data_only=True, read_only=True)
    try:
        rows = list(wb[wb.sheetnames[0]].iter_rows(values_only=True))
    finally:
        wb.close()
    header = [(str(c).strip() if c is not None else "") for c in rows[0]]
    ci = {h: i for i, h in enumerate(header)}
    src_ev = None
    for row in rows[1:]:
        if (
            row[ci["Config Label"]] == config
            and int(round(_to_float(row[ci["n"]]))) == n_expected
        ):
            src_ev = _to_float(row[ci["EV(R)"]])
            break
    assert src_ev is not None, f"{config} (n={n_expected}) row not found in results.xlsx"

    rep_ev = sys.reported_metrics["ev"]
    assert rep_ev == pytest.approx(src_ev, abs=1e-9), (
        f"{name} reported ev={rep_ev!r} vs results.xlsx {config} EV(R)={src_ev!r}"
    )
