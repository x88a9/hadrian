"""Unit tests for the header-driven xlsx parser (no DB, no integration marker).

Synthetic workbooks are built in-memory with openpyxl to cover both real tab
layouts (variant A / variant B), Excel error strings, unfinished backtests and
the direction/win-loss normalization + fallback rules.
"""

from __future__ import annotations

from datetime import datetime

import openpyxl

from app.importers.xlsx import parse_workbook

# Metric header row is identical in both variants; only the value row differs
# (variant A -> row 9, variant B -> row 8).
_METRIC_HEADERS = {
    "B7": "Composite Grade",
    "C7": "EV Grade",
    "D7": "ECE Grade",
    "E7": "EVol Grade",
    "G7": "WIN Rate",
    "H7": "EV",
    "I7": "Total R",
    "J7": "Avg Win R",
    "K7": "Avg Loss R",
    "L7": "Total Trades",
    "M7": "Wins",
    "N7": "Losses",
    "O7": "ECE",
    "P7": "EVol",
}


def _write_metric_headers(ws) -> None:
    for cell, label in _METRIC_HEADERS.items():
        ws[cell] = label


def _write_rules(ws) -> None:
    ws["G3"] = "Entry"
    ws["H3"] = "Break and close of structure"
    ws["G4"] = "Stop Loss"
    ws["H4"] = "Swing low"
    ws["G5"] = "Take Profit"
    ws["H5"] = "3R fixed target"


def _save(wb, tmp_path, name="wb.xlsx") -> str:
    path = tmp_path / name
    wb.save(path)
    return str(path)


def _variant_a_workbook():
    """B2 name, values row 9, Date & Time col C, Direction col J."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "B-H1-801"
    ws["B2"] = "B-H1-801"
    _write_rules(ws)
    _write_metric_headers(ws)
    # value row 9
    ws["B9"] = "C"
    ws["C9"] = "B"
    ws["D9"] = "C"
    ws["E9"] = "B"
    ws["G9"] = 0.33
    ws["H9"] = 0.65
    ws["I9"] = 65.0
    ws["J9"] = 4.0
    ws["K9"] = -1.0
    ws["L9"] = 100
    ws["M9"] = 33
    ws["N9"] = 67
    ws["O9"] = 0.275
    ws["P9"] = 0.193
    # trade header row 13
    for cell, label in {
        "A13": "#",
        "B13": "Day",
        "C13": "Date & Time",
        "D13": "Zone",
        "E13": "Timeframe",
        "F13": "Entry ($)",
        "G13": "Stop Loss ($)",
        "H13": "Exit ($)",
        "J13": "Direction",
        "K13": "R",
        "L13": "W/L",
        "N13": "MA(10)",
    }.items():
        ws[cell] = label
    # two trades
    ws["A14"] = 1
    ws["B14"] = "Friday"
    ws["C14"] = datetime(2025, 10, 3, 12, 0)
    ws["D14"] = "NY AM"
    ws["E14"] = "H1"
    ws["F14"] = 120250.1
    ws["G14"] = 119618.4
    ws["H14"] = 122776.9
    ws["J14"] = "Long"
    ws["K14"] = 4.0
    ws["L14"] = "Win"
    ws["A15"] = 2
    ws["B15"] = "Monday"
    ws["C15"] = datetime(2025, 10, 6, 2, 0)
    ws["E15"] = "H1"
    ws["F15"] = 123427.3
    ws["G15"] = 122552.7
    ws["H15"] = 122552.7
    ws["J15"] = "Short"
    ws["K15"] = -1.0
    ws["L15"] = "Loss"
    return wb


def _variant_b_workbook():
    """B3 name, values row 8, Date Start col C, Direction col I."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TREND-WH4-801"
    ws["B3"] = "TREND-WH4-801"
    _write_rules(ws)
    _write_metric_headers(ws)
    # value row 8
    ws["B8"] = "A"
    ws["C8"] = "A+"
    ws["D8"] = "A"
    ws["E8"] = "C"
    ws["G8"] = 0.4
    ws["H8"] = 9.17
    ws["I8"] = 45.88
    ws["J8"] = 24.44
    ws["K8"] = -1.0
    ws["L8"] = 5
    ws["M8"] = 2
    ws["N8"] = 3
    ws["O8"] = 0.62
    ws["P8"] = 0.14
    # trade header row 13 (variant B: Duration / Date Start / Date End)
    for cell, label in {
        "A13": "#",
        "B13": "Duration",
        "C13": "Date Start",
        "D13": "Date End",
        "E13": "Timeframe",
        "F13": "Entry ($)",
        "G13": "Stop Loss ($)",
        "H13": "Exit ($)",
        "I13": "Direction",
        "J13": "R",
        "K13": "W/L",
    }.items():
        ws[cell] = label
    ws["A14"] = 1
    ws["B14"] = "1w 2d 16h"
    ws["C14"] = datetime(2018, 11, 19, 1, 0)
    ws["D14"] = datetime(2018, 11, 28, 17, 0)
    ws["E14"] = "W / H4"
    ws["F14"] = 5515.09
    ws["G14"] = 5583.04
    ws["H14"] = 4292.47
    ws["I14"] = "Short"
    ws["J14"] = 17.99
    ws["K14"] = "Win"
    return wb


def test_variant_a_layout(tmp_path):
    path = _save(_variant_a_workbook(), tmp_path)
    result = parse_workbook(path)

    assert len(result.tabs) == 1
    tab = result.tabs[0]
    assert tab.tab_name == "B-H1-801"
    assert tab.system_name == "B-H1-801"
    assert tab.parse_status == "complete"

    assert tab.entry_rule == "Break and close of structure"
    assert tab.sl_rule == "Swing low"
    assert tab.tp_rule == "3R fixed target"

    m = tab.reported_metrics
    assert m is not None
    assert m["composite_grade"] == "C"
    assert m["ev_grade"] == "B"
    assert m["ece_grade"] == "C"
    assert m["evol_grade"] == "B"
    assert m["win_rate"] == 0.33
    assert m["ev"] == 0.65
    assert m["total_r"] == 65.0
    assert m["avg_win_r"] == 4.0
    assert m["avg_loss_r"] == -1.0
    assert m["total_trades"] == 100
    assert m["wins"] == 33
    assert m["losses"] == 67
    assert m["ece"] == 0.275
    assert m["evol"] == 0.193

    assert len(tab.trades) == 2
    t0 = tab.trades[0]
    assert t0.number == 1
    assert t0.day == "Friday"
    assert t0.trade_datetime == datetime(2025, 10, 3, 12, 0)
    assert t0.zone == "NY AM"
    assert t0.timeframe == "H1"
    assert t0.entry == 120250.1
    assert t0.sl == 119618.4
    assert t0.exit == 122776.9
    assert t0.direction == "long"
    assert t0.r_value == 4.0
    assert t0.win_loss == "win"

    t1 = tab.trades[1]
    assert t1.direction == "short"
    assert t1.win_loss == "loss"
    assert t1.zone is None


def test_variant_b_layout(tmp_path):
    path = _save(_variant_b_workbook(), tmp_path)
    result = parse_workbook(path)

    tab = result.tabs[0]
    assert tab.system_name == "TREND-WH4-801"  # from B3
    assert tab.parse_status == "complete"

    m = tab.reported_metrics
    assert m is not None
    assert m["composite_grade"] == "A"
    assert m["ev_grade"] == "A+"
    assert m["ev"] == 9.17
    assert m["total_trades"] == 5

    assert len(tab.trades) == 1
    t = tab.trades[0]
    # Date Start is the trade datetime; Date End / Duration ignored.
    assert t.trade_datetime == datetime(2018, 11, 19, 1, 0)
    assert t.direction == "short"
    assert t.r_value == 17.99
    assert t.win_loss == "win"
    assert t.entry == 5515.09


def test_error_strings_become_none_but_status_complete(tmp_path):
    """#DIV/0! in EVol and #N/A in grades -> None; still complete with R values."""
    wb = _variant_a_workbook()
    ws = wb.active
    ws["B9"] = "#N/A"  # composite grade
    ws["D9"] = "#N/A"  # ece grade
    ws["P9"] = "#DIV/0!"  # evol
    ws["O9"] = "#REF!"  # ece
    path = _save(wb, tmp_path)

    tab = parse_workbook(path).tabs[0]
    m = tab.reported_metrics
    assert m["composite_grade"] is None
    assert m["ece_grade"] is None
    assert m["evol"] is None
    assert m["ece"] is None
    # untouched values remain
    assert m["ev"] == 0.65
    # R values still present in the trade log -> complete
    assert tab.parse_status == "complete"


def test_no_numeric_r_is_incomplete(tmp_path):
    """A tab whose trade rows carry no numeric R -> incomplete."""
    wb = _variant_a_workbook()
    ws = wb.active
    ws["K14"] = "#DIV/0!"
    ws["K15"] = "#N/A"
    path = _save(wb, tmp_path)

    tab = parse_workbook(path).tabs[0]
    assert tab.parse_status == "incomplete"
    # trades themselves are still parsed (entry present), R just None
    assert len(tab.trades) == 2
    assert all(t.r_value is None for t in tab.trades)


def test_missing_trade_header_is_skipped(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BROKEN"
    ws["B2"] = "BROKEN-001"
    _write_metric_headers(ws)
    ws["B9"] = "C"
    # no '#' header row anywhere
    path = _save(wb, tmp_path)

    tab = parse_workbook(path).tabs[0]
    assert tab.parse_status == "skipped"
    assert tab.message is not None
    assert tab.system_name == "BROKEN-001"


def test_direction_and_win_loss_normalization_and_fallback(tmp_path):
    """Direction lowercased/validated; W/L falls back to the R rule when blank."""
    wb = _variant_a_workbook()
    ws = wb.active
    # invalid direction -> None; blank W/L -> fallback from R
    ws["J14"] = "sideways"
    ws["L14"] = None  # blank -> fallback: R=4.0 > 0 -> win
    ws["K14"] = 4.0
    # blank W/L, R in the dead band (-0.1 <= R < 0) -> None
    ws["J15"] = "LONG"  # mixed case -> long
    ws["L15"] = None
    ws["K15"] = -0.05
    path = _save(wb, tmp_path)

    tab = parse_workbook(path).tabs[0]
    t0, t1 = tab.trades
    assert t0.direction is None
    assert t0.win_loss == "win"  # fallback from R
    assert t1.direction == "long"
    assert t1.win_loss is None  # dead-band R -> no classification
