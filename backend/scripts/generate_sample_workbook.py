"""Generate the synthetic sample workbook shipped with this repository.

The canonical ``Backtesting Repository.xlsx`` holds real, private research and is
not distributed. This script produces a structurally identical stand-in so that
a fresh clone can exercise the xlsx importer, the metrics layer and the
integration tests without any private data.

Everything in the output is synthetic: system names, rules and trades are
generated from a fixed seed. No figure in it describes a real trading result.

The workbook reproduces every layout feature the parser must cope with
(see ``app/importers/xlsx.py``):

* **Variant A** — system name in ``B2``, metric values in row 9.
* **Variant B** — system name in ``B3``, metric values in row 8, ``Date``
  instead of ``Date & Time`` and Direction/R/W-L shifted one column left.
* An **unfinished tab** whose metric block is full of Excel error sentinels
  (``#DIV/0!``) and which carries no trades, so the importer's skip path stays
  covered.

Usage::

    python backend/scripts/generate_sample_workbook.py [-o OUTPUT]

Re-running it always yields the same trades; the workbook is deterministic.
"""

from __future__ import annotations

import argparse
import random
import sys
from dataclasses import dataclass
from datetime import datetime, timedelta
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

_BACKEND_ROOT = Path(__file__).resolve().parents[1]
if str(_BACKEND_ROOT) not in sys.path:
    sys.path.insert(0, str(_BACKEND_ROOT))

from app.services.metrics import compute_metrics  # noqa: E402

SEED = 20260813
DEFAULT_OUTPUT = (
    _BACKEND_ROOT.parent / "samples" / "backtesting_repository_sample.xlsx"
)

_METRIC_HEADERS = [
    "Composite Grade",
    "EV Grade",
    "ECE Grade",
    "EVol Grade",
    None,
    "WIN Rate",
    "EV",
    "Total R",
    "Avg Win R",
    "Avg Loss R",
    "Total Trades",
    "Wins",
    "Losses",
    "ECE",
    "EVol",
]

_TRADE_HEADERS_A = [
    "#",
    "Day",
    "Date & Time",
    "Zone",
    "Timeframe",
    "Entry ($)",
    "Stop Loss ($)",
    "Exit ($)",
    None,
    "Direction",
    "R",
    "W/L",
]

_TRADE_HEADERS_B = [
    "#",
    "Day",
    "Date",
    "Zone",
    "Timeframe",
    "Entry ($)",
    "Stop Loss ($)",
    "Exit ($)",
    "Direction",
    "R",
    "W/L",
]

_ERROR = "#DIV/0!"


@dataclass
class _Trade:
    """Minimal stand-in satisfying the metrics layer's ``TradeLike`` protocol."""

    number: int
    trade_datetime: datetime
    timeframe: str
    entry: float
    sl: float
    exit: float
    direction: str
    r_value: float
    win_loss: str  # lower-case, as the metrics layer expects
    zone: Optional[str] = None


@dataclass
class _SystemSpec:
    tab: str
    name: str
    timeframe: str
    entry_rule: str
    sl_rule: str
    tp_rule: str
    variant: str
    n_trades: int
    win_rate: float          # in-sample hit rate
    oos_win_rate: float      # out-of-sample hit rate, deliberately lower
    reward: float
    start: datetime
    end: datetime
    base_price: float
    zone: Optional[str] = None


# Identifiers use a 9xx block and generic textbook rules so that nothing here
# collides with, or paraphrases, a real research system.
#
# Two properties are deliberate. First, the set is mixed: one system clears its
# costs, the rest range from marginal to negative. Second, every system performs
# worse out-of-sample than in-sample, which is what real systems do once the
# parameters stop being fitted to the data being measured. A sample without that
# decay would make the IS/OOS split look like a formality.
SPECS: list[_SystemSpec] = [
    _SystemSpec(
        tab="B-H1-901",
        name="B-H1-901",
        timeframe="H1",
        entry_rule="Close above the prior N-bar high",
        sl_rule="Below the signal bar low",
        tp_rule="Fixed multiple of initial risk",
        variant="A",
        n_trades=140,
        win_rate=0.41,
        oos_win_rate=0.27,
        reward=4.0,
        start=datetime(2022, 2, 7, 8, 0),
        end=datetime(2026, 6, 30, 18, 0),
        base_price=32000.0,
    ),
    _SystemSpec(
        tab="TREND-D1-902",
        name="TREND-D1-902",
        timeframe="H1",
        entry_rule="Higher-timeframe trend filter plus a pullback to a moving average",
        sl_rule="Beyond the last opposing swing",
        tp_rule="Exit on moving-average crossover",
        variant="A",
        n_trades=190,
        win_rate=0.47,
        oos_win_rate=0.34,
        reward=2.2,
        start=datetime(2022, 1, 12, 9, 0),
        end=datetime(2026, 7, 15, 14, 0),
        base_price=38000.0,
    ),
    _SystemSpec(
        tab="MR-M15-903",
        name="MR-M15-903",
        timeframe="M15",
        entry_rule="Fade of an N-sigma extension from the session mean",
        sl_rule="One ATR beyond the extension extreme",
        tp_rule="Return to the session mean",
        variant="A",
        n_trades=240,
        win_rate=0.68,
        oos_win_rate=0.50,
        reward=0.85,
        start=datetime(2022, 4, 6, 10, 15),
        end=datetime(2026, 5, 20, 16, 45),
        base_price=27000.0,
    ),
    _SystemSpec(
        tab="REV-H4-904",
        name="REV-H4-904",
        timeframe="H4",
        entry_rule="Reclaim of a range after a failed breakout",
        sl_rule="Beyond the failure extreme",
        tp_rule="Opposite range boundary",
        variant="A",
        n_trades=120,
        win_rate=0.44,
        oos_win_rate=0.30,
        reward=2.6,
        start=datetime(2022, 5, 18, 4, 0),
        end=datetime(2026, 4, 9, 20, 0),
        base_price=24000.0,
    ),
    _SystemSpec(
        tab="BB-M5-905.v2",
        name="BB-M5-905.v2",
        timeframe="M5",
        entry_rule="Band touch followed by a confirming close back inside",
        sl_rule="Beyond the outer band",
        tp_rule="Opposite band",
        variant="B",
        n_trades=88,
        win_rate=0.53,
        oos_win_rate=0.37,
        reward=1.9,
        start=datetime(2022, 3, 9, 0, 0),
        end=datetime(2026, 2, 27, 0, 0),
        base_price=41000.0,
        zone="U.S./New York",
    ),
]

# Unfinished backtest: the importer must skip it and log the reason instead of
# failing the whole run.
UNFINISHED_TAB = "VP-H1-906.wip"


# Trades from this date onward count as out-of-sample; it matches the platform
# default (IS_OOS_SPLIT_DATE).
OOS_START = datetime(2024, 1, 1)


def _r_for(rng: random.Random, win_rate: float, reward: float) -> tuple[float, str]:
    """Draw one trade outcome in R.

    Winners scatter slightly around the nominal target (partial fills, slippage
    against the take-profit), losers sit at -1R with the occasional overshoot
    from a gap through the stop.
    """
    if rng.random() < win_rate:
        return round(reward * rng.uniform(0.94, 1.03), 6), "win"
    if rng.random() < 0.05:  # stop gapped
        return round(-1.0 * rng.uniform(1.05, 1.35), 6), "loss"
    return -1.0, "loss"


def _make_trades(spec: _SystemSpec, rng: random.Random) -> list[_Trade]:
    span = (spec.end - spec.start).total_seconds()
    offsets = sorted(rng.uniform(0, span) for _ in range(spec.n_trades))
    price = spec.base_price
    trades: list[_Trade] = []

    for i, off in enumerate(offsets, start=1):
        ts = spec.start + timedelta(seconds=off)
        # Random walk so prices drift plausibly instead of hovering forever.
        price *= 1.0 + rng.gauss(0, 0.012)
        price = max(price, 1000.0)

        direction = "Long" if rng.random() < 0.55 else "Short"
        risk_pct = rng.uniform(0.004, 0.012)
        entry = round(price, 1)
        # The hit rate degrades once the sample leaves the in-sample period.
        hit_rate = spec.win_rate if ts < OOS_START else spec.oos_win_rate
        r_value, win_loss = _r_for(rng, hit_rate, spec.reward)

        if direction == "Long":
            sl = round(entry * (1 - risk_pct), 1)
            exit_price = round(entry + (entry - sl) * r_value, 1)
        else:
            sl = round(entry * (1 + risk_pct), 1)
            exit_price = round(entry - (sl - entry) * r_value, 1)

        trades.append(
            _Trade(
                number=i,
                trade_datetime=ts.replace(second=0, microsecond=0),
                timeframe=spec.timeframe,
                entry=entry,
                sl=sl,
                exit=exit_price,
                direction=direction,
                r_value=r_value,
                win_loss=win_loss,
                zone=spec.zone,
            )
        )
    return trades


def _write_row(ws: Worksheet, row: int, start_col: int, values: list) -> None:
    for offset, value in enumerate(values):
        if value is None:
            continue
        ws.cell(row=row, column=start_col + offset, value=value)


def _write_header_block(ws: Worksheet, spec: _SystemSpec, metrics: dict) -> None:
    """Name, rule block and reported-metric block, per layout variant."""
    name_row = 2 if spec.variant == "A" else 3
    ws.cell(row=name_row, column=2, value=spec.name).font = Font(bold=True)

    # Rules sit in columns G/H regardless of variant.
    for offset, (label, value) in enumerate(
        (
            ("Entry", spec.entry_rule),
            ("Stop Loss", spec.sl_rule),
            ("Take Profit", spec.tp_rule),
        )
    ):
        ws.cell(row=name_row + offset, column=7, value=label)
        ws.cell(row=name_row + offset, column=8, value=value)

    header_row = 7
    value_row = 9 if spec.variant == "A" else 8
    _write_row(ws, header_row, 2, _METRIC_HEADERS)
    for cell in ws[header_row]:
        if cell.value is not None:
            cell.font = Font(bold=True)

    _write_row(
        ws,
        value_row,
        2,
        [
            metrics["composite_grade"],
            metrics["ev_grade"],
            metrics["ece_grade"],
            metrics["evol_grade"],
            None,
            metrics["win_rate"],
            metrics["ev"],
            metrics["total_r"],
            metrics["avg_win_r"],
            metrics["avg_loss_r"],
            metrics["total_trades"],
            metrics["wins"],
            metrics["losses"],
            metrics["ece"],
            metrics["evol"],
        ],
    )


def _write_trades(ws: Worksheet, spec: _SystemSpec, trades: list[_Trade]) -> None:
    header_row = 13
    headers = _TRADE_HEADERS_A if spec.variant == "A" else _TRADE_HEADERS_B
    _write_row(ws, header_row, 1, headers)
    for cell in ws[header_row]:
        if cell.value is not None:
            cell.font = Font(bold=True)

    for offset, t in enumerate(trades, start=1):
        row = header_row + offset
        if spec.variant == "A":
            values = [
                t.number,
                t.trade_datetime.strftime("%A"),
                t.trade_datetime,
                t.zone,
                t.timeframe,
                t.entry,
                t.sl,
                t.exit,
                None,
                t.direction,
                t.r_value,
                t.win_loss.capitalize(),
            ]
        else:
            values = [
                t.number,
                t.trade_datetime.strftime("%A"),
                t.trade_datetime,
                t.zone,
                t.timeframe,
                t.entry,
                t.sl,
                t.exit,
                t.direction,
                t.r_value,
                t.win_loss.capitalize(),
            ]
        _write_row(ws, row, 1, values)

    for column, width in (
        ("A", 6),
        ("B", 12),
        ("C", 18),
        ("D", 16),
        ("E", 11),
        ("F", 12),
        ("G", 13),
        ("H", 12),
    ):
        ws.column_dimensions[column].width = width


def _write_unfinished(ws: Worksheet) -> None:
    """A tab whose metrics never computed — the shape of a real WIP backtest."""
    ws.cell(row=2, column=2, value=UNFINISHED_TAB).font = Font(bold=True)
    ws.cell(row=2, column=7, value="Entry")
    ws.cell(row=2, column=8, value="Value-area fade on the session profile")
    ws.cell(row=3, column=7, value="Stop Loss")
    ws.cell(row=3, column=8, value="Beyond the value-area extreme")
    ws.cell(row=4, column=7, value="Take Profit")
    ws.cell(row=4, column=8, value="Point of control")

    _write_row(ws, 7, 2, _METRIC_HEADERS)
    _write_row(ws, 9, 2, [_ERROR] * 4 + [None] + [_ERROR] * 8)
    _write_row(ws, 13, 1, _TRADE_HEADERS_A)


def build(output: Path) -> Path:
    rng = random.Random(SEED)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    summary: list[tuple[str, int, Optional[float], Optional[str]]] = []

    for spec in SPECS:
        trades = _make_trades(spec, rng)
        metrics = compute_metrics(trades)
        ws = wb.create_sheet(title=spec.tab)
        _write_header_block(ws, spec, metrics)
        _write_trades(ws, spec, trades)
        summary.append(
            (spec.tab, len(trades), metrics["ev"], metrics["composite_grade"])
        )

    _write_unfinished(wb.create_sheet(title=UNFINISHED_TAB))
    summary.append((UNFINISHED_TAB, 0, None, None))

    output.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output)

    print(f"wrote {output}")
    print(f"{'tab':<18} {'trades':>7} {'EV':>9}  grade")
    for tab, n, ev, grade in summary:
        ev_str = f"{ev:.4f}" if ev is not None else "-"
        print(f"{tab:<18} {n:>7} {ev_str:>9}  {grade or '-'}")
    return output


def main() -> int:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "-o",
        "--output",
        type=Path,
        default=DEFAULT_OUTPUT,
        help=f"target path (default: {DEFAULT_OUTPUT})",
    )
    args = parser.parse_args()
    build(args.output)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
